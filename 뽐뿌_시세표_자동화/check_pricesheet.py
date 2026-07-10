# -*- coding: utf-8 -*-
"""
뽐뿌 카톡 시세표 - 매칭 오류 자동 점검기
사용법:  python check_pricesheet.py "뽐뿌 카톡 시세표.xlsx"
출력  :  같은 폴더에 point_check_report.xlsx  (+ 콘솔 요약)

원리: 출력 시트(3사 통합/SKT/KT/U+)의 각 가격셀 수식(COUNTIFS/SUMIFS)을
      파싱해서 (모델명, 요금제) 키로 ADMIN을 대조.
      - 매칭 1건 = 정상
      - 매칭 0건 = 공백 (모델없음 / 요금제불일치 / 띄어쓰기 불일치 진단)
      - 매칭 2건+ = 중복 → 값 안 뜸
      - COUNTIFS 기준셀 != SUMIFS 기준셀 = 잠재 수식버그
"""
import sys, re, os
import openpyxl
from openpyxl.styles import PatternFill, Font

SRC = sys.argv[1] if len(sys.argv) > 1 else "뽐뿌 카톡 시세표.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(SRC)), "point_check_report.xlsx")

wbf = openpyxl.load_workbook(SRC, data_only=False)
wbv = openpyxl.load_workbook(SRC, data_only=True)
ad  = wbv['ADMIN']

# 통신사별 ADMIN (모델명열, 요금제매칭열)
COLS = {'SKT': ('L', 'P'), 'KT': ('D', 'H'), 'LG': ('T', 'X')}
def cidx(l): return openpyxl.utils.column_index_from_string(l)

admin = {}
for c, (mcol, pcol) in COLS.items():
    mi, pi = cidx(mcol), cidx(pcol)
    rows = []
    for r in range(6, ad.max_row + 1):
        m = ad.cell(r, mi).value; p = ad.cell(r, pi).value
        if m is None and p is None: continue
        rows.append((r, str(m).strip() if m else None, str(p).strip() if p else None))
    admin[c] = rows

norm = lambda s: re.sub(r'\s+', '', s) if s else s
cf_re = re.compile(r'COUNTIFS\((\w+모델명),\$?([A-Z]+)\$?(\d+),(\w+요금제),\$?([A-Z]+)\$?(\d+)\)')
sm_re = re.compile(r'SUMIFS\([^,]+,\w+모델명,\$?[A-Z]+\$?\d+,\w+요금제,\$?([A-Z]+)\$?(\d+)\)')

blanks, dups, planbug = [], [], []
SHEETS = ['3사 통합', 'SKT', 'KT', 'U+']
for sh in SHEETS:
    if sh not in wbf.sheetnames: continue
    fs = wbf[sh]; vs = wbv[sh]
    for row in fs.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or 'COUNTIFS' not in v: continue
            flat = v.replace('\n', '')
            m = cf_re.search(flat)
            if not m: continue
            mrange, mc, mr, prange, pc, pr = m.groups()
            carrier = mrange.replace('모델명', '')
            if carrier not in admin: continue
            model = vs[f'{mc}{mr}'].value; plan = vs[f'{pc}{pr}'].value
            model = str(model).strip() if model else None
            plan  = str(plan).strip()  if plan  else None
            if not model or not plan: continue
            rows = admin[carrier]
            matches = [(r, mm, pp) for (r, mm, pp) in rows if mm == model and pp == plan]
            sm = sm_re.search(flat)
            if sm and (sm.group(1), sm.group(2)) != (pc, pr):
                planbug.append((sh, cell.coordinate, carrier, f'COUNTIFS={pc}{pr} / SUMIFS={sm.group(1)}{sm.group(2)}'))
            if len(matches) == 1: continue
            if len(matches) > 1:
                dups.append((sh, cell.coordinate, carrier, model, plan, len(matches), [r for r,_,_ in matches]))
            else:
                exist = [pp for (r, mm, pp) in rows if mm == model]
                if exist:
                    hint = f'요금제 불일치 (ADMIN엔 {sorted(set(x for x in exist if x))})'
                else:
                    near = [mm for (r, mm, pp) in rows if mm and norm(mm) == norm(model)]
                    hint = f'띄어쓰기 불일치? ADMIN="{near[0]}"' if near else '모델명 ADMIN에 없음(재고없음 가능)'
                blanks.append((sh, cell.coordinate, carrier, model, plan, hint))

# ---- 콘솔 요약 ----
print(f'[중복매칭 → 값 안뜸] {len(dups)}건')
for sh,co,c,mo,pl,n,rs in dups: print(f'  {sh}!{co} {c} | {mo}/{pl} → {n}건 중복 행{rs}')
print(f'[매칭0건 → 공백] {len(blanks)}건')
for sh,co,c,mo,pl,h in blanks: print(f'  {sh}!{co} {c} | {mo}/{pl} → {h}')
print(f'[요금제 기준셀 불일치] {len(planbug)}건')

# ---- xlsx 리포트 ----
wb = openpyxl.Workbook(); ws = wb.active; ws.title = '점검결과'
red = PatternFill('solid', fgColor='FFC7CE'); yel = PatternFill('solid', fgColor='FFEB9C')
gray = PatternFill('solid', fgColor='D9D9D9'); bold = Font(bold=True)
ws.append(['구분','시트','셀','통신사','모델명','요금제','원인/설명']); 
for c in ws[1]: c.font = bold; c.fill = gray
for sh,co,c,mo,pl,n,rs in dups:
    ws.append(['중복매칭',sh,co,c,mo,pl,f'ADMIN {n}건 중복(행 {rs}) → 하나만 남기세요'])
    for cell in ws[ws.max_row]: cell.fill = red
for sh,co,c,mo,pl,h in blanks:
    real = '재고없음' not in h
    ws.append(['매칭0건',sh,co,c,mo,pl,h])
    for cell in ws[ws.max_row]: cell.fill = red if real else yel
for sh,co,c,d in planbug:
    ws.append(['요금제기준셀불일치',sh,co,c,'','',d])
    for cell in ws[ws.max_row]: cell.fill = yel
for col,w in zip('ABCDEFG',[16,10,7,7,22,22,46]): ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
wb.save(OUT)
print(f'\n리포트 저장: {OUT}')
